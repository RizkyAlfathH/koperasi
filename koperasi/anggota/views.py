from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse

from .models import Anggota
from .forms import AnggotaForm, AdminForm
from django.contrib.auth import get_user_model
from django.db.models import Case, When, Value, IntegerField

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from openpyxl import load_workbook
from datetime import date, datetime
import re

User = get_user_model()

# ===============================
# KONFIG ROLE
# ===============================
ROLE_ADMIN = ["admin", "ketua", "sekretaris", "bendahara"]
ROLE_PENGURUS = ["ketua", "sekretaris", "bendahara"]


# ===============================
# DASHBOARD REDIRECT (GLOBAL)
# ===============================
@login_required
def dashboard_redirect(request):
    role = request.user.role

    if role == "ketua":
        return render(request, "dashboard/ketua.html")
    elif role == "sekretaris":
        return render(request, "dashboard/sekretaris.html")
    elif role == "bendahara":
        return render(request, "dashboard/bendahara.html")
    elif role == "admin":
        return redirect("admin_koperasi:admin_dashboard")
    else:
        return redirect("admin_koperasi:admin_login")

# ===============================
# DASHBOARD PER ROLE
# ===============================
@login_required
def ketua_dashboard(request):
    if request.user.role != "ketua":
        return redirect("dashboard")
    return render(request, "dashboard/ketua.html")


@login_required
def sekretaris_dashboard(request):
    if request.user.role != "sekretaris":
        return redirect("dashboard")
    return render(request, "dashboard/sekretaris.html")


@login_required
def bendahara_dashboard(request):
    if request.user.role != "bendahara":
        return redirect("dashboard")
    return render(request, "dashboard/bendahara.html")


# ===============================
# KELOLA AKUN (ROLE-BASED)
# ===============================
@login_required
def kelola_akun(request):
    role = request.user.role   # 🔥 AMBIL DARI LOGIN

    if role not in ROLE_PENGURUS:
        return redirect("dashboard")

    search_admin = request.GET.get("searchAdmin", "")
    search_anggota = request.GET.get("searchAnggota", "")

    # ---------- ADMIN / PENGURUS ----------
    admins = User.objects.filter(role__in=ROLE_PENGURUS)
    if search_admin:
        admins = admins.filter(username__icontains=search_admin)

    paginator_admin = Paginator(admins.order_by("id"), 10)
    admins_page = paginator_admin.get_page(
        request.GET.get("page_admin", 1)
    )

    # ---------- ANGGOTA ----------
    anggotas = Anggota.objects.annotate(
        status_order=Case(
            When(status__iexact="NONAKTIF", then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    )

    if search_anggota:
        anggotas = anggotas.filter(nama__icontains=search_anggota)

    paginator_anggota = Paginator(
        anggotas.order_by("status_order", "nomor_anggota"),
        10
    )
    anggotas_page = paginator_anggota.get_page(
        request.GET.get("page_anggota", 1)
    )

    return render(request, "kelola_akun/kelola_akun.html", {
        "admins": admins_page,
        "anggotas": anggotas_page,
        "searchAdmin": search_admin,
        "searchAnggota": search_anggota,
        "ROLE_ADMIN": ROLE_ADMIN,
    })

# ===============================
# CRUD ADMIN (KETUA / SEKRETARIS / BENDAHARA)
# ===============================
@login_required
def tambah_admin(request):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    if request.method == "POST":
        form = AdminForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Admin berhasil ditambahkan.")
            return redirect("kelola_akun", role=request.user.role)
    else:
        form = AdminForm()

    return render(request, "kelola_akun/Form/form_admin.html", {
        "form": form,
        "judul": "Tambah Admin"
    })


@login_required
def edit_admin(request, user_id):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    admin = get_object_or_404(User, id=user_id, role__in=ROLE_PENGURUS)
    form = AdminForm(request.POST or None, instance=admin)

    if form.is_valid():
        form.save()
        messages.success(request, "Admin berhasil diperbarui.")
        return redirect("kelola_akun")

    return render(request, "kelola_akun/Form/form_admin.html", {
        "form": form,
        "judul": "Edit Admin"
    })


@login_required
def hapus_admin(request, user_id):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    admin = get_object_or_404(User, id=user_id, role__in=ROLE_PENGURUS)
    admin.delete()
    messages.success(request, "Admin berhasil dihapus.")
    return redirect("kelola_akun")

def detail_admin(request, user_id):
    admin = get_object_or_404(User, id=user_id)

    nomor_urut = User.objects.filter(id__lte=admin.id).count()

    context = {
        "admin": admin,
        "nomor_urut": nomor_urut,
    }
    return render(request, "kelola_akun/detail/detail_admin.html", context)


# ===============================
# CRUD ANGGOTA
# ===============================
@login_required
def tambah_anggota(request):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    if request.method == "POST":
        form = AnggotaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Anggota berhasil ditambahkan.")
            return redirect("kelola_akun")
    else:
        form = AnggotaForm()

    return render(request, "kelola_akun/Form/form_anggota.html", {
        "form": form,
        "judul": "Tambah Anggota"
    })


@login_required
def edit_anggota(request, nomor_anggota):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    anggota = get_object_or_404(Anggota, nomor_anggota=nomor_anggota)
    form = AnggotaForm(request.POST or None, instance=anggota)

    if form.is_valid():
        form.save()
        messages.success(request, "Anggota berhasil diperbarui.")
        return redirect("kelola_akun")

    return render(request, "kelola_akun/Form/form_anggota.html", {
        "form": form,
        "judul": "Edit Anggota"
    })


@login_required
def hapus_anggota(request, nomor_anggota):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    anggota = get_object_or_404(Anggota, nomor_anggota=nomor_anggota)
    anggota.delete()
    messages.success(request, "Anggota berhasil dihapus.")
    return redirect("kelola_akun")

def detail_anggota(request, nomor_anggota):
    anggota = get_object_or_404(Anggota, nomor_anggota=nomor_anggota)

    context = {
        "anggota": anggota,
    }
    return render(request, "kelola_akun/detail/detail_anggota.html", context)


# ===============================
# API VALIDASI
# ===============================
@login_required
def cek_email(request):
    email = request.GET.get("email", "")
    return JsonResponse({
        "exists": Anggota.objects.filter(email=email).exists()
    })

# ===============================
# EXPORT EXCEL DATA ANGGOTA
# ===============================

@login_required
def export_excel_anggota(request):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Anggota"

    headers = [
        "No. Anggota", "Nama", "NIP", "Alamat", "No. Telepon",
        "Email", "Jenis Kelamin", "Tanggal Daftar",
        "Status", "Tanggal Nonaktif", "Alasan Nonaktif"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="FFFF00")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for anggota in Anggota.objects.all().order_by("nomor_anggota"):
        ws.append([
            anggota.nomor_anggota,
            anggota.nama,
            anggota.nip,
            anggota.alamat,
            anggota.no_telp,
            anggota.email,
            anggota.jenis_kelamin,
            anggota.tanggal_daftar.strftime("%d-%m-%Y") if anggota.tanggal_daftar else "-",
            anggota.status,
            anggota.tanggal_nonaktif.strftime("%d-%m-%Y") if anggota.tanggal_nonaktif else "-",
            anggota.alasan_nonaktif or "-"
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="anggota_koperasi.xlsx"'
    wb.save(response)
    return response

# ===============================
# EXPORT PDF DATA ANGGOTA
# ===============================

@login_required
def export_pdf_anggota(request):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="anggota_koperasi.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    normal.fontSize = 7

    data = [[
        "No. Anggota", "Nama", "NIP", "Alamat", "No. Telp",
        "Email", "JK", "Tgl Daftar", "Status", "Tgl Nonaktif", "Alasan"
    ]]

    for a in Anggota.objects.all().order_by("nomor_anggota"):
        data.append([
            a.nomor_anggota,
            Paragraph(a.nama or "-", normal),
            a.nip or "-",
            Paragraph(a.alamat or "-", normal),
            a.no_telp or "-",
            Paragraph(a.email or "-", normal),
            a.jenis_kelamin,
            a.tanggal_daftar.strftime("%d-%m-%Y") if a.tanggal_daftar else "-",
            a.status,
            a.tanggal_nonaktif.strftime("%d-%m-%Y") if a.tanggal_nonaktif else "-",
            Paragraph(a.alasan_nonaktif or "-", normal),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.yellow),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 7),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))

    doc.build([table])
    return response


# ===============================
# IMPORT EXCEL DATA ANGGOTA
# ===============================

@login_required
def import_excel_anggota(request):
    if request.user.role not in ROLE_ADMIN:
        return redirect("dashboard")

    if request.method == "POST" and request.FILES.get("excel_file"):
        try:
            wb = load_workbook(request.FILES["excel_file"], data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "File Excel tidak bisa dibaca")
            return redirect("kelola_akun")

        sukses = 0
        gagal = 0

        # mulai dari baris ke-4 (header sampai baris 3)
        for idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
            try:
                # ===== NOMOR ANGGOTA =====
                nomor_anggota = str(row[1].value).strip() if row[1].value else None
                nama = row[2].value

                if not nomor_anggota or not nama:
                    continue

                # hanya format NA xxx
                if not re.match(r"^NA\s*\d+", nomor_anggota):
                    continue

                # ===== JENIS KELAMIN =====
                jk_excel = str(row[4].value).strip().upper() if row[4].value else ""

                if jk_excel == "L":
                    jenis_kelamin = "Laki-laki"
                elif jk_excel == "P":
                    jenis_kelamin = "Perempuan"
                else:
                    jenis_kelamin = "Laki-laki"

                # ===== UMUR =====
                umur = row[3].value if isinstance(row[3].value, int) else None

                # ===== PEKERJAAN =====
                pekerjaan = row[5].value or "-"

                # ===== ALAMAT =====
                alamat = row[6].value or "-"

                # ===== TANGGAL DAFTAR =====
                tgl_daftar = row[8].value
                if isinstance(tgl_daftar, datetime):
                    tgl_daftar = tgl_daftar.date()
                else:
                    tgl_daftar = date.today()

                # ===== TANGGAL NONAKTIF =====
                tgl_nonaktif = row[12].value
                if isinstance(tgl_nonaktif, datetime):
                    tgl_nonaktif = tgl_nonaktif.date()
                else:
                    tgl_nonaktif = None

                # ===== ALASAN NONAKTIF =====
                alasan_nonaktif = row[13].value or "-"

                # ===== SIMPAN / UPDATE =====
                anggota, created = Anggota.objects.update_or_create(
                    nomor_anggota=nomor_anggota,
                    defaults={
                        "nama": nama,
                        "umur": umur,
                        "jenis_kelamin": jenis_kelamin,
                        "pekerjaan": pekerjaan,
                        "alamat": alamat,
                        "tanggal_daftar": tgl_daftar,
                        "tanggal_nonaktif": tgl_nonaktif,
                        "status": "nonaktif" if tgl_nonaktif else "aktif",
                        "alasan_nonaktif": alasan_nonaktif,

                        # kolom yang tidak ada di Excel
                        "nip": "-",
                        "no_telp": "-",
                        "email": "-",
                    }
                )

                # 🔐 PASSWORD DEFAULT (HANYA JIKA BARU)
                if created:
                    anggota.set_password("12345")
                    anggota.save()

                sukses += 1

            except Exception as e:
                gagal += 1
                # print(f"Error baris {idx}: {e}")

        messages.success(
            request,
            f"Import selesai: {sukses} berhasil, {gagal} gagal"
        )
        return redirect("kelola_akun")

    messages.error(request, "File Excel tidak valid")
    return redirect("kelola_akun")